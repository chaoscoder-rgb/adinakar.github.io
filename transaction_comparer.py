import pandas as pd
import ast

def load_and_preprocess_data():
    """
    Loads data from CSV files into pandas DataFrames and performs initial preprocessing.
    Expected files:
    - 'files/AD-TransactionAnalysis-Fields.csv'
    - 'files/splunk_log_data_pre.csv'
    - 'files/splunk_log_data_post.csv'
    """
    try:
        # Load the analysis fields mapping
        df_analysis_fields = pd.read_csv('files/AD-TransactionAnalysis-Fields.csv')

        # Preprocess df_analysis_fields
        # Convert 'Weight' from "10%" string to float 0.10
        df_analysis_fields['Weight_Num'] = df_analysis_fields['Weight'].str.rstrip('%').astype('float') / 100.0
        # Strip potential leading/trailing whitespace from column names and values
        df_analysis_fields.columns = df_analysis_fields.columns.str.strip()
        for col in df_analysis_fields.columns:
            if df_analysis_fields[col].dtype == 'object':
                df_analysis_fields[col] = df_analysis_fields[col].str.strip()

        # Load the pre and post log data
        df_pre_logs = pd.read_csv('files/splunk_log_data_pre.csv')
        df_post_logs = pd.read_csv('files/splunk_log_data_post.csv')

        # Strip potential leading/trailing whitespace from column names
        df_pre_logs.columns = df_pre_logs.columns.str.strip()
        df_post_logs.columns = df_post_logs.columns.str.strip()

        # Identify columns that need to be parsed from string-formatted dicts
        # Based on AD-TransactionAnalysis-Fields.csv, these are typically:
        # Request Headers, Request Payload, Response Body
        # However, it's safer to check which parameters have examples suggesting dict structures.
        dict_like_params = []
        if 'Example' in df_analysis_fields.columns:
            for _, row in df_analysis_fields.iterrows():
                example_val = str(row['Example'])
                if example_val.startswith('{') and example_val.endswith('}'):
                    dict_like_params.append(row['Parameter'])

        # Ensure we only try to parse columns that actually exist in the log DataFrames
        cols_to_parse_pre = [col for col in dict_like_params if col in df_pre_logs.columns]
        cols_to_parse_post = [col for col in dict_like_params if col in df_post_logs.columns]

        # Helper function to safely parse string-formatted dictionaries
        def parse_string_dict(s):
            if pd.isna(s) or s == "" or not isinstance(s, str):
                return None # Or {} if an empty dict is preferred for missing values
            try:
                return ast.literal_eval(s)
            except (ValueError, SyntaxError):
                # Handle cases where parsing fails, e.g., malformed string
                # print(f"Warning: Could not parse string: {s}")
                return s # Keep original if parsing fails, or return a specific error marker

        for col in cols_to_parse_pre:
            df_pre_logs[col] = df_pre_logs[col].apply(parse_string_dict)

        for col in cols_to_parse_post:
            df_post_logs[col] = df_post_logs[col].apply(parse_string_dict)

        # Convert relevant columns to string to ensure consistent comparisons,
        # especially for numerical-looking codes that should be treated as strings.
        # This should be done after parsing for dict-like columns.
        string_cols = ['User ID', 'Requesting Service', 'Error Code', 'HTTP Method',
                       'Responding Service', 'Release Version', 'Location', 'Device/OS', 'Session ID']

        for df in [df_pre_logs, df_post_logs]:
            for col in string_cols:
                if col in df.columns:
                    df[col] = df[col].astype(str).fillna('') # Convert to string and fill NaN with empty string

        # Error Description might also need to be string, but it's handled by default read_csv as object/string
        if 'Error Description' in df_pre_logs.columns:
             df_pre_logs['Error Description'] = df_pre_logs['Error Description'].astype(str).fillna('')
        if 'Error Description' in df_post_logs.columns:
             df_post_logs['Error Description'] = df_post_logs['Error Description'].astype(str).fillna('')


        print("Data loaded and preprocessed successfully.")
        print("\nAnalysis Fields Head:")
        print(df_analysis_fields.head())
        print(f"\nAnalysis Fields Columns: {df_analysis_fields.columns.tolist()}")
        print(f"Identified dict-like parameters for parsing: {dict_like_params}")


        print("\nPre-Logs Head:")
        print(df_pre_logs.head())
        print(f"\nPre-Logs Columns: {df_pre_logs.columns.tolist()}")
        if cols_to_parse_pre:
            print(f"\nSample of parsed '{cols_to_parse_pre[0]}' in Pre-Logs:")
            print(df_pre_logs[cols_to_parse_pre[0]].head())


        print("\nPost-Logs Head:")
        print(df_post_logs.head())
        print(f"\nPost-Logs Columns: {df_post_logs.columns.tolist()}")
        if cols_to_parse_post:
            print(f"\nSample of parsed '{cols_to_parse_post[0]}' in Post-Logs:")
            print(df_post_logs[cols_to_parse_post[0]].head())

        return df_analysis_fields, df_pre_logs, df_post_logs
    except Exception as e:
        print(f"An error occurred during data loading and preprocessing: {e}")
        # Optionally re-raise the exception if you want the program to stop
        # raise
        # Or return None or empty DataFrames to indicate failure
        return None, None, None

if __name__ == '__main__':
    # This is for testing the function directly
    df_analysis, df_pre, df_post = load_and_preprocess_data()

    if df_analysis is None or df_pre is None or df_post is None:
        print("Data loading failed. Exiting.")
    else:
        # Further checks can be added here, e.g.
        print("\nData types in df_analysis_fields:")
    print(df_analysis.dtypes)
    print("\nData types in df_pre_logs:")
    print(df_pre.dtypes)
    print("\nData types in df_post_logs:")
    print(df_post.dtypes)

    # Example: Check one of the parsed columns
    if 'Request Headers' in df_post.columns:
        print("\nExample of a parsed Request Header entry in df_post_logs:")
        print(df_post['Request Headers'].iloc[0])
        print(type(df_post['Request Headers'].iloc[0]))

    if 'Weight_Num' in df_analysis.columns:
        print("\nNumeric weights in df_analysis_fields:")
        print(df_analysis[['Parameter', 'Weight', 'Weight_Num']].head())

    # Check for NaNs in key fields after processing
    baseline_fields = df_analysis[df_analysis['Match Group'] == 'Baseline']['Parameter'].tolist()
    print(f"\nBaseline fields: {baseline_fields}")

    print("\nNaN counts in baseline fields of Pre-Logs:")
    for col in baseline_fields:
        if col in df_pre.columns:
            print(f"{col}: {df_pre[col].isna().sum()}")
        else:
            print(f"{col}: Not found in Pre-Logs")

    print("\nNaN counts in baseline fields of Post-Logs:")
    for col in baseline_fields:
        if col in df_post.columns:
            print(f"{col}: {df_post[col].isna().sum()}")
        else:
            print(f"{col}: Not found in Post-Logs")

    # Check Requesting Service example from AD-TransactionAnalysis-Fields.csv
    # "payment-service (From headers)", Service name in X-Service-Name
    # The current logic directly uses the "Requesting Service" column.
    # If X-Service-Name is a header in 'Request Headers', that would require different logic.
    # For now, assuming direct match on 'Requesting Service' column.
    if 'Requesting Service' in df_post.columns:
        print("\nSample 'Requesting Service' data from post_logs:")
        print(df_post['Requesting Service'].head())

    if 'Error Description' in df_post.columns:
        print("\nSample 'Error Description' data from post_logs:")
        print(df_post['Error Description'].head())
        print(type(df_post['Error Description'].iloc[0]))


    # Check if 'Error Code' is string
    if 'Error Code' in df_post.columns:
        print("\nType of 'Error Code' in post_logs")
        print(df_post['Error Code'].dtype)
        print(df_post['Error Code'].head())

    # Check a string-parsed dict column to ensure it's dict
    if 'Request Payload' in df_post and not df_post['Request Payload'].empty:
        first_payload_item = df_post['Request Payload'].iloc[0]
        print(f"\nFirst item in 'Request Payload' (post_logs): {first_payload_item}")
        print(f"Type of first item in 'Request Payload' (post_logs): {type(first_payload_item)}")

    if 'Response Body' in df_post and not df_post['Response Body'].empty:
        first_response_body_item = df_post['Response Body'].iloc[0]
        print(f"\nFirst item in 'Response Body' (post_logs): {first_response_body_item}")
        print(f"Type of first item in 'Response Body' (post_logs): {type(first_response_body_item)}")

    print("Test run of load_and_preprocess_data finished.")


def compare_transactions(df_post_logs, df_pre_logs, df_analysis_fields):
    """
    Compares transactions from df_post_logs with df_pre_logs based on df_analysis_fields.

    Args:
        df_post_logs (pd.DataFrame): DataFrame of current logs.
        df_pre_logs (pd.DataFrame): DataFrame of historical/baseline logs.
        df_analysis_fields (pd.DataFrame): DataFrame with matching rules and weights.

    Returns:
        pd.DataFrame: df_post_logs with added 'Match Result' and 'Total Matched Percentage' columns.
    """
    baseline_fields = df_analysis_fields[df_analysis_fields['Match Group'] == 'Baseline']['Parameter'].tolist()
    comprehensive_fields = df_analysis_fields[df_analysis_fields['Match Group'] == 'Comprehensive']['Parameter'].tolist()

    all_matchable_fields = baseline_fields + comprehensive_fields

    # Create a dictionary for quick weight lookup
    # Ensure 'Parameter' is the index for quick lookup
    if 'Parameter' not in df_analysis_fields.columns:
        raise ValueError("Critical column 'Parameter' not found in df_analysis_fields.")
    if 'Weight_Num' not in df_analysis_fields.columns:
        # Attempt to create it if 'Weight' exists
        if 'Weight' in df_analysis_fields.columns:
            print("Warning: 'Weight_Num' not found, attempting to create from 'Weight'.")
            df_analysis_fields['Weight_Num'] = df_analysis_fields['Weight'].str.rstrip('%').astype('float') / 100.0
        else:
            raise ValueError("Critical column 'Weight_Num' (and 'Weight') not found in df_analysis_fields.")

    weights_map = df_analysis_fields.set_index('Parameter')['Weight_Num'].to_dict()

    results = [] # Will store dicts for each post_row

    for index, post_row in df_post_logs.iterrows():
        match_details_for_row = {
            'Match Result': 'No',
            'Total Matched Percentage': 'Not Applicable',
            'Matched Pre Row Index': None, # Store index of the matched pre_row
            'Field Matches': {} # Store individual field match status (True/False)
        }

        # Temp variable to store the actual pre_row data if a match is found
        # This won't be directly in the main output CSV/Excel rows but used for styling
        matched_pre_row_data = None

        for pre_index, pre_row in df_pre_logs.iterrows():
            baseline_match = True
            current_field_matches_for_pre_row = {} # For this specific pre_row comparison

            # Check baseline fields first
            for field in baseline_fields:
                if field not in post_row or field not in pre_row:
                    baseline_match = False
                    break
                post_val = post_row[field]
                pre_val = pre_row[field]

                is_field_match = False
                if pd.isna(post_val) and pd.isna(pre_val):
                    is_field_match = True
                elif not pd.isna(post_val) and not pd.isna(pre_val) and post_val == pre_val:
                    is_field_match = True

                current_field_matches_for_pre_row[field] = is_field_match
                if not is_field_match:
                    baseline_match = False
                    # break # Don't break here, continue checking other baseline fields for this pre_row
                           # to capture all baseline mismatches if this pre_row was a candidate

            if not baseline_match: # If any baseline field did not match for this pre_row
                # Reset current_field_matches_for_pre_row or ensure it reflects only this attempt
                # This is important if we were to pick the 'best' pre_row match.
                # For now, we take the first baseline match.
                continue # Try next pre_row

            # If we reach here, all baseline fields matched for the current pre_row
            match_details_for_row['Match Result'] = 'Yes'
            match_details_for_row['Matched Pre Row Index'] = pre_index
            matched_pre_row_data = pre_row # Store this specific pre_row

            current_transaction_match_percentage = 0.0

            # Now, evaluate all fields (baseline + comprehensive) for percentage and detailed field matching
            # against this specific matched_pre_row_data
            for field in all_matchable_fields:
                field_match_status = False
                if field in post_row and field in matched_pre_row_data:
                    post_val = post_row[field]
                    pre_val = matched_pre_row_data[field]

                    if pd.isna(post_val) and pd.isna(pre_val):
                        field_match_status = True
                    elif not pd.isna(post_val) and not pd.isna(pre_val) and post_val == pre_val:
                        field_match_status = True

                    match_details_for_row['Field Matches'][field] = field_match_status
                    if field_match_status:
                        current_transaction_match_percentage += weights_map.get(field, 0)
                else:
                    match_details_for_row['Field Matches'][field] = False # Field not present in one or both

            match_details_for_row['Total Matched Percentage'] = round(current_transaction_match_percentage * 100, 2)
            break # Found a baseline match and processed it, move to the next post_row

        results.append(match_details_for_row)

    # Construct the output DataFrame
    output_df = df_post_logs.copy()

    # Add 'Match Result' and 'Total Matched Percentage' from the results list
    output_df['Match Result'] = [res['Match Result'] for res in results]
    output_df['Total Matched Percentage'] = [res['Total Matched Percentage'] for res in results]

    # Store the detailed field matches and matched pre_row index for later use (styling)
    # These won't be columns in the final Excel but attributes or separate structures
    # For simplicity in this step, we can return these details along with the output_df
    # The main script can then pass them to the Excel writing function.

    # For now, the function signature returns only the output_df.
    # We need to decide how to pass 'results' (which contains Field Matches and Matched Pre Row Index)
    # to the styling function.
    # Option 1: Return a tuple: (output_df, results_details_list)
    # Option 2: Add them as temporary columns to output_df and drop before final save (less clean)

    # Let's go with Option 1 for clarity.
    return output_df, results


if __name__ == '__main__':
    # This is for testing the function directly
    df_analysis, df_pre, df_post = load_and_preprocess_data()

    # Further checks can be added here, e.g.
    print("\nData types in df_analysis_fields:")
    print(df_analysis.dtypes)
    # print("\nData types in df_pre_logs:")
    # print(df_pre.dtypes)
    # print("\nData types in df_post_logs:")
    # print(df_post.dtypes)

    if 'Weight_Num' in df_analysis.columns:
        print("\nNumeric weights in df_analysis_fields:")
        print(df_analysis[['Parameter', 'Weight', 'Weight_Num']].head())

    print("\nRunning transaction comparison...")
    # Expecting a tuple: (DataFrame, list_of_match_details_dicts)
    comparison_output_df, detailed_results = compare_transactions(df_post, df_pre, df_analysis)

    print("\nComparison Output DF Head:")
    print(comparison_output_df.head())

    print("\nValue counts for 'Match Result':")
    print(comparison_output_df['Match Result'].value_counts())

    print("\nSample of 'Total Matched Percentage' for 'Yes' matches:")
    print(comparison_output_df[comparison_output_df['Match Result'] == 'Yes']['Total Matched Percentage'].head())

    print("\nSample of 'Total Matched Percentage' for 'No' matches:")
    print(comparison_output_df[comparison_output_df['Match Result'] == 'No']['Total Matched Percentage'].head())

    # Example: Print details for the first 'Yes' match found
    first_yes_match_index = -1
    for i, detail in enumerate(detailed_results):
        if detail['Match Result'] == 'Yes':
            first_yes_match_index = i
            break

    if first_yes_match_index != -1:
        print(f"\nDetails for first 'Yes' match (post_row index {first_yes_match_index}):")
        print(f"  Matched Pre Row Index: {detailed_results[first_yes_match_index].get('Matched Pre Row Index')}")
        print(f"  Field Matches: {detailed_results[first_yes_match_index].get('Field Matches')}")


    # # Save the output to an Excel file for inspection (styling will be separate)
    # try:
    #     output_file_path_excel = 'files/comparison_output.xlsx'
    #     # At this stage, we save the df without styling. Styling will be a separate step.
    #     comparison_output_df.to_excel(output_file_path_excel, index=False, engine='openpyxl')
    #     print(f"\nComparison output (without styling) saved to {output_file_path_excel}")
    # except Exception as e:
    #     print(f"Error saving output to Excel: {e}")

    # print("Test run of compare_transactions (modified) finished.")

# --- Excel Writing with Styling ---
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def write_excel_output_with_styling(output_df, detailed_results, df_analysis_fields,
                                   pre_file_name, post_file_name, num_pre_rows, num_post_rows, # For summary sheet later
                                   filename="files/comparison_output.xlsx"): # Updated filename
    """
    Writes the comparison output to an Excel file with conditional styling on the main sheet
    and adds a summary sheet.
    """
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Transaction Comparison"

    # Define fills
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid") # Light Green
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")   # Light Red

    # Prepare a copy of the DataFrame for writing to Excel, converting dicts/lists to strings
    df_for_excel = output_df.copy()
    for col in df_for_excel.columns:
        # Check if the column is of object type and its first non-NA element is a dict or list
        if df_for_excel[col].dtype == 'object':
            first_valid_item = df_for_excel[col].dropna().iloc[0] if not df_for_excel[col].dropna().empty else None
            if isinstance(first_valid_item, (dict, list)):
                df_for_excel[col] = df_for_excel[col].apply(lambda x: str(x) if isinstance(x, (dict, list)) else x)

    # Append DataFrame to the main worksheet
    for r_idx, row_data in enumerate(dataframe_to_rows(df_for_excel, index=False, header=True), 1):
        ws_main.append(row_data)

    # Get the list of all columns that could be styled
    # These are the original columns from post_logs that are also in analysis_fields
    # and are present in the output_df
    styleable_columns = [param for param in df_analysis_fields['Parameter'].tolist() if param in output_df.columns]

    header_list = [cell.value for cell in ws_main[1]] # Get header names from the sheet
    col_name_to_idx = {name: i + 1 for i, name in enumerate(header_list)}


    # Apply styling (starting from row 2, as row 1 is header)
    for ws_row_index, detailed_info in enumerate(detailed_results, 2): # ws_row_index is 2-based for sheet rows
        if detailed_info['Match Result'] == 'Yes':
            field_matches = detailed_info.get('Field Matches', {})
            for field_name, is_match in field_matches.items():
                if field_name in col_name_to_idx: # Ensure the field is a column in the Excel sheet
                    col_idx = col_name_to_idx[field_name]
                    cell = ws_main.cell(row=ws_row_index, column=col_idx)
                    if is_match:
                        cell.fill = green_fill
                    else:
                        # Only apply red if it's a field that's supposed to be matched.
                        # This prevents coloring non-data columns if they somehow end up in field_matches.
                        if field_name in styleable_columns:
                             cell.fill = red_fill

    # --- Summary Sheet Creation ---
    ws_summary = wb.create_sheet("Comparison Summary")

    # Add static info and overall stats
    summary_data = [
        ("Pre-File Name:", pre_file_name),
        ("Post-File Name:", post_file_name),
        ("Number of Rows in Pre-File:", num_pre_rows),
        ("Number of Rows in Post-File:", num_post_rows),
        ("Count of Exact Matches (Baseline):", output_df[output_df['Match Result'] == 'Yes'].shape[0]),
        ("Count of Mismatches (Baseline):", output_df[output_df['Match Result'] == 'No'].shape[0]),
    ]

    for r_idx, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=r_idx, column=1, value=label)
        ws_summary.cell(row=r_idx, column=2, value=value)

    # Add a small gap before the pivot table
    current_row_in_summary = len(summary_data) + 2
    ws_summary.cell(row=current_row_in_summary, column=1, value="Matches Grouped by User ID, Error Code, Requesting Service:")
    current_row_in_summary += 1

    # Generate and write pivot table
    if not output_df[output_df['Match Result'] == 'Yes'].empty:
        matched_df = output_df[output_df['Match Result'] == 'Yes']
        try:
            pivot_table = pd.pivot_table(
                matched_df,
                index=['User ID', 'Error Code', 'Requesting Service'],
                aggfunc='size', # Count occurrences
                fill_value=0
            ).reset_index(name='Match Count') # Convert to DataFrame and name the count column

            # Write pivot table header
            for c_idx, value in enumerate(pivot_table.columns, 1):
                ws_summary.cell(row=current_row_in_summary, column=c_idx, value=value)
            current_row_in_summary +=1

            # Write pivot table rows
            for _, pivot_row in pivot_table.iterrows():
                for c_idx, value in enumerate(pivot_row, 1):
                    ws_summary.cell(row=current_row_in_summary, column=c_idx, value=value)
                current_row_in_summary += 1
        except Exception as e:
            ws_summary.cell(row=current_row_in_summary, column=1, value=f"Error generating pivot table: {e}")
    else:
        ws_summary.cell(row=current_row_in_summary, column=1, value="No 'Yes' matches to generate pivot table.")

    try:
        wb.save(filename)
        print(f"\nStyled comparison output with summary sheet saved to {filename}")
    except Exception as e:
        print(f"Error saving styled Excel output: {e}")


if __name__ == '__main__':
    # This is for testing the function directly
    df_analysis, df_pre, df_post = load_and_preprocess_data()

    if df_analysis is None or df_pre is None or df_post is None:
        print("Data loading failed. Exiting.")
    else:
        print("\nData types in df_analysis_fields:")
        print(df_analysis.dtypes)

        if 'Weight_Num' in df_analysis.columns:
            print("\nNumeric weights in df_analysis_fields:")
            print(df_analysis[['Parameter', 'Weight', 'Weight_Num']].head())

        print("\nRunning transaction comparison...")
        comparison_output_df, detailed_results = compare_transactions(df_post, df_pre, df_analysis)

        print("\nComparison Output DF Head:")
        print(comparison_output_df.head())

        print("\nValue counts for 'Match Result':")
        print(comparison_output_df['Match Result'].value_counts())
        print("\nSample of 'Total Matched Percentage' for 'Yes' matches:")
        print(comparison_output_df[comparison_output_df['Match Result'] == 'Yes']['Total Matched Percentage'].head())
        print("\nSample of 'Total Matched Percentage' for 'No' matches:")
        print(comparison_output_df[comparison_output_df['Match Result'] == 'No']['Total Matched Percentage'].head())

        first_yes_match_index = -1
        for i, detail in enumerate(detailed_results):
            if detail['Match Result'] == 'Yes':
                first_yes_match_index = i
                break
        if first_yes_match_index != -1:
            print(f"\nDetails for first 'Yes' match (post_row index {first_yes_match_index}):")
            print(f"  Matched Pre Row Index: {detailed_results[first_yes_match_index].get('Matched Pre Row Index')}")
            # Limiting print output for long Field Matches dictionary
            field_matches_sample = dict(list(detailed_results[first_yes_match_index].get('Field Matches', {}).items())[:5])
            print(f"  Field Matches (sample): {field_matches_sample}")


        # Call the new function to write styled Excel output
        # File names and row counts for summary sheet (will be used in next step)
        pre_file_name = "splunk_log_data_pre.csv" # Example, ideally get from actual load
        post_file_name = "splunk_log_data_post.csv"
        num_pre_rows = len(df_pre)
        num_post_rows = len(df_post)

        write_excel_output_with_styling(
            comparison_output_df,
            detailed_results,
            df_analysis, # Pass df_analysis_fields for column names
            pre_file_name, post_file_name, num_pre_rows, num_post_rows, # Pass summary info
            filename="files/comparison_output.xlsx" # Ensure it's .xlsx
        )

        print("Test run with styled Excel output finished.")
